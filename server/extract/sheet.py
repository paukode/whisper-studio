"""Spreadsheet extraction.

Small sheets keep MarkItDown's full Markdown table. Once that output gets large
(a wide sheet with thousands of rows easily exceeds the context window), we swap
in a compact schema + sample instead: each sheet's dimensions, header row, and
first N rows. The model can then ask for specific rows/columns via the
analyze_document tool.
"""

import io
import logging

log = logging.getLogger("whisper-studio")

# MarkItDown output above this many chars -> switch to schema + sample.
_LARGE_CHARS = 60_000
_SAMPLE_ROWS = 20


def extract_xlsx(content: bytes, markitdown_text: str) -> str:
    if len(markitdown_text or "") <= _LARGE_CHARS:
        return markitdown_text
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        log.warning("openpyxl load failed (%s); keeping markitdown output", e)
        return markitdown_text

    parts = [
        "[Large spreadsheet: showing each sheet's dimensions, header, and first "
        f"{_SAMPLE_ROWS} rows. Ask for specific rows or columns to see more.]"
    ]
    try:
        for ws in wb.worksheets:
            parts.append(_summarize_ws(ws))
    finally:
        wb.close()
    return "\n\n".join(parts)


def extract_csv(content: bytes, markitdown_text: str) -> str:
    if len(markitdown_text or "") <= _LARGE_CHARS:
        return markitdown_text
    try:
        import csv

        text = content.decode("utf-8", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
    except Exception as e:
        log.warning("csv parse failed (%s); keeping markitdown output", e)
        return markitdown_text
    if not rows:
        return markitdown_text

    header = rows[0]
    sample = rows[1 : _SAMPLE_ROWS + 1]
    total = len(rows) - 1
    note = (
        f"[Large CSV: {total} data rows x {len(header)} cols; showing the first "
        f"{_SAMPLE_ROWS}. Ask for specific rows or columns to see more.]"
    )
    return note + "\n\n" + _rows_to_md(header, sample)


def _summarize_ws(ws) -> str:
    rows = ws.max_row or 0
    cols = ws.max_column or 0
    header: list[str] = []
    sample: list[list[str]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = ["" if c is None else str(c) for c in row]
            continue
        if i > _SAMPLE_ROWS:
            break
        sample.append(["" if c is None else str(c) for c in row])

    title = f"### Sheet: {ws.title}  ({rows} rows x {cols} cols)"
    if not header:
        return title + "\n(empty)"
    return title + "\n" + _rows_to_md(header, sample)


def _rows_to_md(header, rows) -> str:
    def esc(v) -> str:
        return str(v).replace("|", "\\|").replace("\n", " ")

    out = [
        "| " + " | ".join(esc(h) for h in header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for r in rows:
        cells = list(r) + [""] * (len(header) - len(r))
        out.append("| " + " | ".join(esc(c) for c in cells[: len(header)]) + " |")
    return "\n".join(out)
