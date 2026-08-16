"""The app's standard document layout, applied by every document tool.

One place defines what "a normal-looking document" means so create_docx /
create_pptx / create_xlsx / create_pdf all produce consistent, professional
output by default, instead of each underlying library's own quirky default
(python-docx: US Letter, 11pt Calibri body but blue Calibri Light headings;
reportlab: US Letter 10pt; python-pptx: the ancient 4:3 slide size).

The standard:
- docx: A4 page, 1-inch margins, Calibri 11pt body, 16/13/12pt headings.
- pptx: 16:9 widescreen (PowerPoint's own modern default), Calibri theme.
- xlsx: Calibri 11 (openpyxl default), bold header row, frozen top row,
  column widths fitted to content.
- pdf: A4 page, 1-inch margins, 11pt body with comfortable leading.
"""

# ── docx ──────────────────────────────────────────────────────────────

# A4 in millimetres. python-docx's bundled template is US Letter.
DOCX_PAGE_WIDTH_MM = 210
DOCX_PAGE_HEIGHT_MM = 297
DOCX_MARGIN_INCHES = 1

DOCX_BODY_FONT = "Calibri"
DOCX_BODY_SIZE_PT = 11
# h1/h2/h3 carry the house sizes; h4-h6 inherit the template's own ramp.
DOCX_HEADING_SIZES_PT = {1: 16, 2: 13, 3: 12}


def apply_docx_standard(doc) -> None:
    """Put an open python-docx Document on the house standard: A4 page,
    1-inch margins, Calibri 11pt body, black 16/13/12pt headings.

    Set on the STYLES, so it holds for content added afterwards and for
    anything the renderer does not stamp per run. Headings need the explicit
    pass because the bundled template styles them as blue Calibri Light,
    which reads as a template rather than as a plain document."""
    from docx.oxml.ns import qn
    from docx.shared import Inches, Mm, Pt, RGBColor

    for section in doc.sections:
        section.page_width = Mm(DOCX_PAGE_WIDTH_MM)
        section.page_height = Mm(DOCX_PAGE_HEIGHT_MM)
        section.left_margin = Inches(DOCX_MARGIN_INCHES)
        section.right_margin = Inches(DOCX_MARGIN_INCHES)
        section.top_margin = Inches(DOCX_MARGIN_INCHES)
        section.bottom_margin = Inches(DOCX_MARGIN_INCHES)

    normal = doc.styles["Normal"]
    normal.font.name = DOCX_BODY_FONT
    normal.font.size = Pt(DOCX_BODY_SIZE_PT)
    # font.name only writes w:ascii/w:hAnsi; without the east-asian slot Word
    # substitutes its own font for any run it decides is not Latin.
    normal.element.rPr.rFonts.set(qn("w:eastAsia"), DOCX_BODY_FONT)

    for level in range(1, 7):
        try:
            style = doc.styles[f"Heading {level}"]
        except KeyError:
            continue
        style.font.name = DOCX_BODY_FONT
        style.font.bold = True
        style.font.color.rgb = RGBColor(0, 0, 0)
        size = DOCX_HEADING_SIZES_PT.get(level)
        if size:
            style.font.size = Pt(size)


# ── pptx ──────────────────────────────────────────────────────────────

# PowerPoint's own 16:9 dimensions in EMU (13.333 x 7.5 inches).
PPTX_SLIDE_WIDTH_EMU = 12192000
PPTX_SLIDE_HEIGHT_EMU = 6858000


def apply_pptx_standard(prs) -> None:
    """Switch a python-pptx Presentation to 16:9 — its bundled template
    still defaults to 4:3. The template's Office theme already uses
    Calibri, so fonts need no per-run overrides."""
    prs.slide_width = PPTX_SLIDE_WIDTH_EMU
    prs.slide_height = PPTX_SLIDE_HEIGHT_EMU


def fit_pptx_placeholders(prs, slide) -> None:
    """Stretch a freshly-added slide's placeholders across the 16:9 width.
    python-pptx copies placeholder geometry from the 4:3 layout, which
    would leave a third of every widescreen slide empty on the right."""
    for shape in slide.placeholders:
        shape.width = prs.slide_width - 2 * shape.left


# ── xlsx ──────────────────────────────────────────────────────────────

XLSX_MIN_COL_WIDTH = 10
XLSX_MAX_COL_WIDTH = 60


def apply_xlsx_standard(worksheet) -> None:
    """Standard worksheet polish: bold header row, frozen top row, and
    column widths fitted to content (bounded so one long cell can't blow a
    column out to absurd width). openpyxl's defaults already give Calibri
    11 for cell text."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"
    for idx, column in enumerate(worksheet.iter_cols(), start=1):
        longest = max((len(str(c.value)) for c in column if c.value is not None), default=0)
        width = min(max(XLSX_MIN_COL_WIDTH, longest + 2), XLSX_MAX_COL_WIDTH)
        worksheet.column_dimensions[get_column_letter(idx)].width = width


# ── pdf ───────────────────────────────────────────────────────────────

PDF_BODY_FONT_SIZE = 11
PDF_BODY_LEADING = 15


def pdf_standard_styles():
    """reportlab stylesheet adjusted to the standard: 11pt body with
    comfortable leading (the sample sheet's Normal is a cramped 10/12).
    Callers pair this with an A4 SimpleDocTemplate."""
    from reportlab.lib.styles import getSampleStyleSheet

    styles = getSampleStyleSheet()
    styles["Normal"].fontSize = PDF_BODY_FONT_SIZE
    styles["Normal"].leading = PDF_BODY_LEADING
    return styles
