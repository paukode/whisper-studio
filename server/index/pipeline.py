"""Indexing orchestration + retrieval.

``build()`` walks a workspace, decides what changed (size/mtime gate → content
hash), and only re-embeds new/changed files; deletes drop missing files. The
embedder and GLiNER are unloaded afterward so an index run doesn't keep ~1 GB
resident. ``query()`` is the retrieval side: embed the question, vector-search,
then take one GraphRAG hop along shared entities.
"""

from __future__ import annotations

import hashlib
import logging
import os
import re
import threading
import time
from datetime import datetime, timezone

from . import (
    contextualize,
    descriptions,
    embedder,
    extractor,
    paths,
    relations,
    relstore,
    salience,
    store,
    wssettings,
)
from .chunker import chunk_text, section_path
from .citations import citation_link
from .config import (
    COHERE_EMBED_MODEL_ID,
    DATA_EXTENSIONS,
    EMBED_MODEL,
    GROUND_REL_FLOOR,
    GROUND_SCORE_FLOORS,
    IMAGE_EXTENSIONS,
    MAX_FILE_BYTES,
    MEDIA_EXTENSIONS,
    MEDIA_MAX_FILE_BYTES,
    RERANK_CANDIDATES,
    RICH_DOC_EXTENSIONS,
    RICH_MAX_FILE_BYTES,
    SHEET_EXTENSIONS,
    SHEET_MAX_FILE_BYTES,
    TEXT_EXTENSIONS,
    labels_for_profile,
    profile_for_ext,
)

log = logging.getLogger("whisper-studio")

# Directories never worth indexing — VCS, deps, build output, caches.
_IGNORE_DIRS = {
    ".git",
    ".svn",
    ".hg",
    "node_modules",
    "venv",
    ".venv",
    "__pycache__",
    ".mypy_cache",
    ".pytest_cache",
    "dist",
    "build",
    ".next",
    ".cache",
    "target",
    ".idea",
    ".vscode",
    "models",
    ".DS_Store",
}

# Everything we know how to turn into text: plain text/code, spreadsheets
# (schema+sample), rich docs (MarkItDown + PDF OCR), images (OCR), and
# audio/video (local transcription).
_INDEXABLE = (
    TEXT_EXTENSIONS | SHEET_EXTENSIONS | RICH_DOC_EXTENSIONS | IMAGE_EXTENSIONS | MEDIA_EXTENSIONS
)

# Heading-structured docs (markdown + rich docs MarkItDown renders with # headings):
# a chunk's section breadcrumb is folded into its embedding input. Excludes code
# (a Python "# comment" is not a heading) and data files.
_HEADING_DOC_EXTS = {".md", ".mdx", ".rst"} | RICH_DOC_EXTENSIONS

# One lock per workspace so a scheduled refresh and a user-triggered build never
# run concurrently on the same index and race the SQLite write transactions.
_build_locks: dict[str, threading.Lock] = {}
_build_locks_guard = threading.Lock()

# Cooperative cancel: a per-workspace Event the build loop checks between files.
_cancel_events: dict[str, threading.Event] = {}
_cancel_guard = threading.Lock()


# Base64 image blobs MarkItDown inlines for DOCX/HTML are huge embedding noise.
_DATA_URI_RE = re.compile(r"data:[\w.+-]+/[\w.+-]+;base64,[A-Za-z0-9+/=]+")


def _ws_lock(ws_root: str) -> threading.Lock:
    with _build_locks_guard:
        lk = _build_locks.get(ws_root)
        if lk is None:
            lk = threading.Lock()
            _build_locks[ws_root] = lk
        return lk


def _cancel_event(ws_root: str) -> threading.Event:
    with _cancel_guard:
        ev = _cancel_events.get(ws_root)
        if ev is None:
            ev = threading.Event()
            _cancel_events[ws_root] = ev
        return ev


def is_building(ws_path: str) -> bool:
    """True if a build currently holds this workspace's lock (e.g. the daily
    scheduled refresh is running), so callers can avoid queuing behind it."""
    ws_root = os.path.abspath(os.path.expanduser(ws_path))
    lk = _build_locks.get(ws_root)
    return bool(lk and lk.locked())


def request_cancel(ws_path: str) -> bool:
    """Ask an in-progress build for this workspace to stop after the current
    file. Returns True if a build was running to signal."""
    if not is_building(ws_path):
        return False
    ws_root = os.path.abspath(os.path.expanduser(ws_path))
    _cancel_event(ws_root).set()
    return True


def _iter_files(ws_root: str):
    """Yield (abs_path, rel_path) for every indexable file under ws_root."""
    for dirpath, dirnames, filenames in os.walk(ws_root):
        dirnames[:] = [d for d in dirnames if d not in _IGNORE_DIRS and not d.startswith(".")]
        for name in filenames:
            ext = os.path.splitext(name)[1].lower()
            if ext not in _INDEXABLE:
                continue
            ap = os.path.join(dirpath, name)
            yield ap, os.path.relpath(ap, ws_root)


def _file_hash(path: str) -> str:
    h = hashlib.sha1()
    with open(path, "rb") as f:
        for block in iter(lambda: f.read(65536), b""):
            h.update(block)
    return h.hexdigest()


def _read_text(path: str) -> str | None:
    try:
        with open(path, "rb") as f:
            raw = f.read(MAX_FILE_BYTES + 1)
    except OSError:
        return None
    if len(raw) > MAX_FILE_BYTES or b"\x00" in raw:
        return None  # too big, or binary
    try:
        return raw.decode("utf-8")
    except UnicodeDecodeError:
        return raw.decode("utf-8", errors="ignore")


_SHEET_SAMPLE_ROWS = 20  # rows kept when sampling a large spreadsheet


def _finish(text: str | None) -> str | None:
    """Normalize extractor output: strip + drop empty / 'nothing here' sentinels
    (server/extract emits '[No content extracted]' / '[No text recognized]')."""
    text = (text or "").strip()
    if not text or text in {"[No content extracted]", "[No text recognized]"}:
        return None
    return text


def _read_bytes(path: str) -> bytes | None:
    try:
        with open(path, "rb") as f:
            return f.read()
    except OSError:
        return None


def _rows_to_md(header, rows) -> str:
    head = "| " + " | ".join(str(h) for h in header) + " |"
    sep = "| " + " | ".join("---" for _ in header) + " |"
    body = "\n".join("| " + " | ".join(str(c) for c in r) + " |" for r in rows)
    return "\n".join([head, sep, body])


def _sample_large_sheet(path: str, ext: str) -> str | None:
    """Bounded schema + sample for a big spreadsheet — streamed, never loading
    the whole grid into memory and never invoking MarkItDown."""
    try:
        if ext == ".csv":
            import csv

            sample_rows: list = []
            extra = 0
            with open(path, encoding="utf-8", errors="ignore", newline="") as f:
                reader = csv.reader(f)
                for row in reader:
                    if len(sample_rows) < _SHEET_SAMPLE_ROWS + 1:
                        sample_rows.append(row)
                    else:
                        extra += 1
            if not sample_rows:
                return None
            header, sample = sample_rows[0], sample_rows[1:]
            total = len(sample) + extra
            note = (
                f"[Large CSV: {total} data rows x {len(header)} columns; "
                f"showing the first {len(sample)}.]"
            )
            return note + "\n\n" + _rows_to_md(header, sample)
        # xlsx/xls — openpyxl read-only streams rows without loading the workbook.
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        parts: list[str] = []
        try:
            for ws in wb.worksheets:
                collected: list = []
                for r in ws.iter_rows(values_only=True):
                    collected.append(r)
                    if len(collected) >= _SHEET_SAMPLE_ROWS + 1:
                        break
                if not collected:
                    continue
                header = ["" if c is None else str(c) for c in collected[0]]
                sample = [["" if c is None else str(c) for c in r] for r in collected[1:]]
                parts.append(
                    f"## {ws.title}\n[~{ws.max_row or '?'} rows x "
                    f"{ws.max_column or len(header)} cols; first {len(sample)}]\n\n"
                    + _rows_to_md(header, sample)
                )
        finally:
            wb.close()
        return "\n\n".join(parts) or None
    except Exception as e:  # noqa: BLE001
        log.warning("Large-sheet sampling failed for %s: %s", path, e)
        return None


def _convert_doc(path: str, ext: str, base: str) -> str | None:
    raw = _read_bytes(path)
    if raw is None:
        return None
    try:
        from server.extract import convert_document

        return _finish(_DATA_URI_RE.sub("[image]", convert_document(raw, ext, base) or ""))
    except Exception as e:  # noqa: BLE001 — one bad file must not abort the build
        log.warning("Index extraction failed for %s: %s", base, e)
        return None


def _ocr_doc(path: str, base: str) -> str | None:
    raw = _read_bytes(path)
    if raw is None:
        return None
    try:
        from server.extract import ocr_image_bytes

        return _finish(ocr_image_bytes(raw))
    except Exception as e:  # noqa: BLE001
        log.warning("Index OCR failed for %s: %s", base, e)
        return None


def _media_doc(path: str, ext: str, base: str) -> str | None:
    raw = _read_bytes(path)
    if raw is None:
        return None
    try:
        from server.extract.media import extract_media

        text, _frames = extract_media(raw, ext, base)  # index wants text only
        return _finish(text)
    except Exception as e:  # noqa: BLE001
        log.warning("Index media extraction failed for %s: %s", base, e)
        return None


def _extract_text(path: str, ext: str, rel: str) -> str | None:
    """Turn any indexable file into text, by type.

    - text/code → decode directly.
    - spreadsheets → small ones via convert_document; large ones via a bounded,
      streamed schema+sample (no full load, no MarkItDown).
    - rich docs (pdf/docx/pptx/html/epub…) → convert_document (MarkItDown + PDF OCR).
    - images → OCR (Apple Vision → Haiku).
    - audio/video → read into memory then transcribed locally.
    Returns None when nothing usable comes out (binary, empty, oversized, failed).
    """
    if ext in TEXT_EXTENSIONS:
        return _read_text(path)
    try:
        size = os.path.getsize(path)
    except OSError:
        return None
    base = os.path.basename(rel)

    if ext in SHEET_EXTENSIONS:
        if size > SHEET_MAX_FILE_BYTES:
            log.info("Index: skipping oversized sheet (%d bytes): %s", size, rel)
            return None
        if size > MAX_FILE_BYTES:  # large → streamed sample
            return _finish(_sample_large_sheet(path, ext))
        if ext == ".csv":  # small CSV → fast direct read
            return _read_text(path)
        return _convert_doc(path, ext, base)  # small xlsx/xls
    if ext in MEDIA_EXTENSIONS:
        if size > MEDIA_MAX_FILE_BYTES:
            log.info("Index: skipping oversized media (%d bytes): %s", size, rel)
            return None
        return _media_doc(path, ext, base)
    if ext in IMAGE_EXTENSIONS:
        if size > RICH_MAX_FILE_BYTES:
            log.info("Index: skipping oversized image (%d bytes): %s", size, rel)
            return None
        return _ocr_doc(path, base)
    # Other rich docs (pdf/docx/pptx/html/epub).
    if size > RICH_MAX_FILE_BYTES:
        log.info("Index: skipping oversized file (%d bytes): %s", size, rel)
        return None
    return _convert_doc(path, ext, base)


def build(ws_path: str, progress=None) -> dict:
    """Incrementally (re)index ``ws_path``. Returns final stats.

    ``progress(done, total, current_path)`` is called as files are processed.
    Serialized per workspace so a scheduled refresh and a manual build can't
    race the same index.
    """
    ws_root = os.path.abspath(os.path.expanduser(ws_path))
    with _ws_lock(ws_root):
        return _build_locked(ws_path, ws_root, progress)


def _build_locked(ws_path: str, ws_root: str, progress) -> dict:
    manifest = store.get_manifest(ws_path)
    # Record the workspace path up front so even an interrupted build is
    # discoverable by list_indexed_workspaces / the scheduler / migration.
    # last_indexed_at is only written on completion (below) — an index with no
    # last_indexed_at is one that never finished.
    store.set_meta(ws_path, workspace=ws_root)
    cancel = _cancel_event(ws_root)
    cancel.clear()
    candidates = list(_iter_files(ws_root))
    total = len(candidates)
    seen: set[str] = set()
    changed = 0
    cancelled = False
    # Read this workspace's typed-relations setting once per build (enabled +
    # which engine: cloud Haiku or on-device Gemma).
    _ws_settings = wssettings.get_settings(ws_path)
    _tr = _ws_settings["typed_relations"]
    typed_rel = _tr["enabled"]
    typed_engine = _tr["engine"]
    _desc = _ws_settings["entity_descriptions"]
    # Contextual chunk headers prepended before embedding (point: metadata in the
    # embedding). A change to the mode (e.g. off -> filename) must re-embed even
    # unchanged files, so detect it against the mode used at the last build.
    _cc = _ws_settings["chunk_context"]
    cc_mode = _cc["mode"]
    cc_engine = _cc["engine"]
    # Which on-device NER model runs (gliner default / gliner2). The entity label
    # set (business docs vs source code) is picked per file from its extension
    # inside the loop below, so mixed folders tag each file with the right labels.
    ner_model = _ws_settings.get("ner_model", "gliner")
    # Treat a pre-feature index (no stamp) as "off": its chunks were embedded
    # header-less, so explicitly choosing filename/llm MUST re-embed. (The earlier
    # "grandfather None → never force" logic silently skipped a user's explicit
    # switch, stamping the new mode while leaving the old header-less vectors.)
    _prev_cc = store.get_meta(ws_path).get("context_mode") or "off"
    force_ctx_reembed = _prev_cc != cc_mode

    try:
        for i, (ap, rel) in enumerate(candidates):
            if cancel.is_set():
                cancelled = True
                break
            seen.add(rel)
            if progress:
                progress(i, total, rel)
            try:
                st = os.stat(ap)
            except OSError:
                continue
            prev = manifest.get(rel)
            # Cheap gate: unchanged size + mtime → skip without hashing. Bypassed
            # when the contextual-header mode changed (every file must re-embed).
            if (
                not force_ctx_reembed
                and prev
                and prev["size"] == st.st_size
                and prev["mtime"] == st.st_mtime
            ):
                continue
            h = _file_hash(ap)
            if not force_ctx_reembed and prev and prev["hash"] == h:
                store.touch_file(ws_path, rel, st.st_size, st.st_mtime)
                continue
            ext = os.path.splitext(rel)[1].lower()
            # Auto-pick the entity label set from the file type: source code gets
            # the code labels, everything else the business set (no user choice).
            entity_labels = labels_for_profile(profile_for_ext(ext))
            fmeta = {"hash": h, "size": st.st_size, "mtime": st.st_mtime}
            text = _extract_text(ap, ext, rel)
            chunks = chunk_text(text) if text is not None else []
            if not chunks:
                # Record an empty manifest entry (0 chunks) so the cheap gate
                # skips this file next run instead of re-extracting/re-OCR'ing
                # it on every daily refresh forever.
                store.replace_file(ws_path, rel, fmeta, [])
                continue
            ctxs = _chunk_contexts(ws_root, rel, text, chunks, cc_mode, cc_engine)
            # For heading-structured docs, fold each chunk's section breadcrumb into
            # its embedding input (embedding-only — stored text and line anchors are
            # untouched), so a section carries its document context ("Overview >
            # Q3 results") into the vector.
            heading_doc = ext in _HEADING_DOC_EXTS
            embed_inputs = []
            for cx, c in zip(ctxs, chunks, strict=False):
                bc = section_path(text, c["start_line"]) if heading_doc else ""
                header = "\n".join(part for part in (cx, f"Section: {bc}" if bc else "") if part)
                embed_inputs.append(f"{header}\n\n{c['text']}" if header else c["text"])
            vecs = embedder.embed_documents(embed_inputs)
            # Structured-data files stay searchable but are not entity-mined:
            # their keys are schema, not graph entities, and are the top junk
            # source. (Content is still embedded + BM25-indexed below.)
            skip_ner = ext in DATA_EXTENSIONS
            records = []
            for c, v in zip(chunks, vecs, strict=False):
                records.append(
                    {
                        "start_line": c["start_line"],
                        "end_line": c["end_line"],
                        "text": c["text"],
                        "vec": v,
                        "entities": (
                            []
                            if skip_ner
                            else extractor.extract_entities(
                                c["text"], labels=entity_labels, ner_model=ner_model
                            )
                        ),
                    }
                )
            store.replace_file(ws_path, rel, fmeta, records)
            if typed_rel:
                names = [e["name"] for rec in records for e in rec.get("entities", [])]
                # The relation engine is chosen independently of the entity model:
                # "gliner2" extracts relations natively (one local model pass, no
                # LLM); "haiku"/"local" use the LLM engine. So a workspace can pair
                # GLiNER2 entities with an LLM's richer relations, or go fully local.
                if typed_engine == "gliner2":
                    rels = extractor.extract_relations_gliner2(text, names, labels=entity_labels)
                else:
                    rels = relations.extract_relations(text, names, typed_engine)
                store.set_file_relations(ws_path, rel, rels)
                # Node-id-keyed mirror (survives entity merges, assistant-queryable
                # via workspace_graph_query). Attach a verbatim evidence line so
                # each fact cites the exact lines via the #wsfile &L anchor.
                facts = []
                for s, t, ty, sc in rels:
                    sl, el, ev = relstore.evidence_line(text, s, t)
                    facts.append(
                        {
                            "source": s,
                            "target": t,
                            "predicate": ty,
                            "strength": sc,
                            "evidence": ev,
                            "start_line": sl,
                            "end_line": el,
                        }
                    )
                relstore.set_file_relations_v2(ws_path, rel, facts)
            changed += 1

        # On cancel we stopped mid-walk, so `seen` is incomplete — skip the
        # deletion pass (it would wrongly drop unvisited files) and the
        # completion stamp (the index is partial). Whatever was indexed stays.
        if not cancelled:
            for gone in set(manifest) - seen:
                store.delete_file(ws_path, gone)
            # Collapse case/spacing/punctuation variants of the same entity into
            # one node so the graph shows one bubble per person/topic.
            store.dedupe_entities(ws_path)
            # Score every entity for salience (name shape x GLiNER confidence x
            # IDF, with a boilerplate-hub cap) so junk is downweighted in the
            # graph hop and graph views instead of hard-deleted.
            salience.recompute(ws_path)
            # Optional: one LLM description per canonical entity (off by default).
            # Runs after dedup so each canonical node is described exactly once.
            if _desc["enabled"]:
                _items = store.entities_for_description(ws_path)
                store.set_node_descriptions(
                    ws_path, descriptions.describe_entities(_items, _desc["engine"])
                )
            from server.infrastructure.model_mode import resolve_backend

            _embed_backend = resolve_backend("embed")
            # Stamp the ACTUAL embed model for the backend that built this index,
            # not a hardcoded Qwen3 id: a Cohere-built index must record the Cohere
            # model, else meta and the stats "embed_model" misreport how the stored
            # vectors were produced.
            embed_model = COHERE_EMBED_MODEL_ID if _embed_backend == "cohere" else EMBED_MODEL
            store.set_meta(
                ws_path,
                last_indexed_at=datetime.now(timezone.utc).isoformat(),
                embed_model=embed_model,
                embed_backend=_embed_backend,
                workspace=ws_root,
                context_mode=cc_mode,
            )
    finally:
        cancel.clear()
        # Free the heavy models — an index run shouldn't hold RAM afterward.
        embedder.unload()
        extractor.unload()
        from . import reranker

        reranker.unload()  # no-op if a chat turn never warmed it

    s = store.stats(ws_path)
    s["changed_files"] = changed
    s["removed_files"] = 0 if cancelled else len(set(manifest) - seen)
    s["cancelled"] = cancelled
    log.info(
        "Index build for %s: %d files, %d chunks (%d changed)%s",
        ws_root,
        s["files"],
        s["chunks"],
        changed,
        " [cancelled]" if cancelled else "",
    )
    return s


def _chunk_contexts(
    ws_root: str, rel: str, doc_text: str, chunks: list[dict], mode: str, engine: str
) -> list[str]:
    """One context header per chunk, per the folder's chunk_context mode:
    - "off"      → no header (embed bare content)
    - "filename" → the file path (free, offline): puts "buddy" from buddy.json
                   into the embedding even when the content never says it
    - "llm"      → an LLM situating line per chunk (contextualize module)
    """
    header = f"File: {os.path.basename(ws_root)}/{rel}"
    if mode == "filename":
        return [header] * len(chunks)
    if mode == "llm":
        # Fall back to the filename header for any chunk the LLM couldn't
        # contextualize (offline, model not downloaded, parse failure) — so "llm"
        # is never WORSE than "filename", just richer when it works.
        llm = contextualize.contextualize_chunks(
            f"{os.path.basename(ws_root)}/{rel}", doc_text, [c["text"] for c in chunks], engine
        )
        return [c or header for c in llm]
    return [""] * len(chunks)


def query(ws_path: str, text: str, k: int = 8, graph_hop: bool = True) -> dict:
    """Retrieve for a question: dense vector top-k, the BM25 keyword top-k, and
    one GraphRAG hop — returned SEPARATELY so the caller fuses them.

    Returns ``{matches, keyword, related}``: ``matches`` are dense (cosine-
    ranked) chunks, ``keyword`` are BM25 chunks (each with ``_bm25``; gated by
    the ``rag_hybrid_search`` flag) catching exact terms — filenames, ids, codes
    — that dense embeddings miss, and ``related`` are the entity-graph hop off
    the dense matches. They're kept apart so a multi-index caller can fuse the
    dense lists globally (cosine is comparable) rather than per-index by rank.
    """
    from server.infrastructure.feature_flags import is_enabled

    t0 = time.time()
    qv = embedder.embed_query(text)
    matches = store.search(ws_path, qv, k=k)
    keyword: list[dict] = []
    if is_enabled("rag_hybrid_search"):
        try:
            keyword = store.fts_search(ws_path, text, k=k)
        except Exception as e:  # noqa: BLE001 — keyword leg is best-effort
            log.warning("Keyword (BM25) search failed for %s: %s", ws_path, e)
    related = []
    if graph_hop and matches:
        related = store.expand(ws_path, [m["chunk_id"] for m in matches], limit=max(2, k // 2))
    log.info(
        "Index query %r → %d dense, %d keyword, %d related in %.2fs",
        text[:60],
        len(matches),
        len(keyword),
        len(related),
        time.time() - t0,
    )
    return {"matches": matches, "keyword": keyword, "related": related}


# Grounding excerpt budget. Matches get a near-full chunk (chunks are bounded
# at ~CHUNK_MAX_TOKENS, ≈1.5K chars) so the model can answer straight from the
# index instead of re-opening files; the graph-hop "related" passages are
# supporting context, so they get a shorter excerpt.
_GROUND_MATCH_CHARS = 1500
_GROUND_RELATED_CHARS = 600
# Drop vector matches below this cosine score: they're noise that only dilutes
# the grounding block (and tempts the model to distrust the passages).
_GROUND_SCORE_FLOOR = 0.15
# Max chunks any single file may contribute to the grounding block (matches AND
# graph-hop related combined), so one large document (e.g. a 300-chunk book)
# can't monopolise the slots and crowd out the small-but-relevant file (e.g. a
# 7-chunk CV). The cross-index candidate pool backfills the freed slots.
_GROUND_PER_DOC_CAP = 3
# Auto-merge (parent-document retrieval): once a document has a chunk in the top
# matches, pull in its OTHER chunks so a small, focused file grounds in full
# (e.g. a CV's complete employment history, not just its header). Gated so a
# large document (a 300-chunk book) is NEVER whole-merged, and bounded by a
# global budget so a few small docs can't blow up the context either.
_AUTOMERGE_MAX_DOC_CHUNKS = 12  # only whole-merge docs with <= this many chunks
_AUTOMERGE_BUDGET = 10  # max sibling chunks added across all docs


def message_text(m: dict) -> str:
    """Plain text of a chat message whose ``content`` may be a string OR a list
    of content blocks (tool_use / tool_result / image, as the backend stores
    multi-turn tool history). Returns only the text-block text; non-text content
    contributes nothing rather than crashing a naive ``.strip()``."""
    c = m.get("content")
    if isinstance(c, str):
        return c
    if isinstance(c, list):
        return " ".join(
            b.get("text", "") for b in c if isinstance(b, dict) and b.get("type") == "text"
        )
    return ""


def build_context_query(question: str, history: list[dict] | None, max_chars: int = 600) -> str:
    """Tier 1 contextualization: fold the most recent turn(s) into the retrieval
    query so a follow-up like "what about Ada?" still carries the prior intent
    ("jobs and durations"). Pure string work — no LLM, no added latency. Returns
    ``question`` unchanged when there's no usable history."""
    if not history:
        return question
    recent: list[str] = []
    for m in reversed(history):
        if m.get("role") in ("user", "assistant"):
            content = message_text(m).strip()
            if content:
                recent.append(content[:300])
        if len(recent) >= 2:
            break
    if not recent:
        return question
    recent.reverse()
    # Prior turns as supporting context, current question last (the dominant
    # signal), the whole thing capped so context can't drown the question.
    return f"{' '.join(recent)[:max_chars]} {question}".strip()


def _rrf_fuse(ranked_lists: list[list[dict]], rrf_k: int = 60) -> list[dict]:
    """Reciprocal-rank fusion of several cosine-ranked match lists into one
    ranking. Rank-based, so it's robust to the different score scales of two
    queries (raw vs contextualized); each chunk keeps its best cosine as the
    displayed score and is ordered by summed reciprocal rank."""
    fused: dict[tuple, dict] = {}
    for lst in ranked_lists:
        for rank, m in enumerate(lst, start=1):
            key = (m.get("_abs"), m.get("start_line"), m.get("end_line"))
            entry = fused.get(key)
            contrib = 1.0 / (rrf_k + rank)
            if entry is None:
                fused[key] = {"m": m, "rrf": contrib, "cos": float(m.get("score", 0.0))}
            else:
                entry["rrf"] += contrib
                entry["cos"] = max(entry["cos"], float(m.get("score", 0.0)))
    out: list[dict] = []
    for entry in sorted(fused.values(), key=lambda e: e["rrf"], reverse=True):
        m = dict(entry["m"])
        m["score"] = entry["cos"]
        out.append(m)
    return out


def retrieve_grounding(
    index_paths: list[str],
    question: str,
    k: int = 12,
    *,
    extra_queries: list[str] | None = None,
    return_meta: bool = False,
):
    """Search the given indexed workspaces and return a cited Markdown context
    block (or '' if nothing). Used to ground a chat answer when the user has
    selected indexes to search.

    Retrieves for ``question`` plus any ``extra_queries`` (e.g. a
    conversation-contextualized variant) and fuses the rankings with reciprocal
    rank fusion, so a follow-up recalls the right passages without one noisy
    query dominating. Includes the top vector matches AND the GraphRAG hop
    (chunks linked through shared entities), with near-full-chunk excerpts, and
    instructs the model to answer from these passages rather than re-reading the
    files. Source links use absolute paths so they reveal in Finder regardless
    of which workspace is connected."""
    queries = [question]
    for q in extra_queries or []:
        if q and q.strip() and q not in queries:
            queries.append(q)

    # Resolve the distinct, indexed roots once (shared by every query).
    seen_roots: set[str] = set()
    valid_roots: list[tuple[str, str]] = []
    for p in index_paths:
        if not p:
            continue
        root = os.path.normpath(os.path.abspath(os.path.expanduser(p)))
        if root in seen_roots or not paths.is_indexed(p):
            continue
        seen_roots.add(root)
        valid_roots.append((p, root))

    # Per-backend cosine floor — score distributions differ across embedders.
    try:
        from server.infrastructure.model_mode import resolve_backend

        _embed_backend = resolve_backend("embed")
    except Exception:
        _embed_backend = "qwen3"
    ground_floor = GROUND_SCORE_FLOORS.get(_embed_backend, _GROUND_SCORE_FLOOR)

    # Dedup on the FULL normalized text (not a prefix): a duplicate file is
    # byte-identical so it collapses, while two distinct invoices that merely
    # share a letterhead/header differ later in the body and are both kept.
    def _text_key(m: dict) -> str:
        return " ".join((m.get("text") or "").split()).lower()

    def _key(m: dict) -> tuple:
        return (m.get("_abs"), m.get("start_line"), m.get("end_line"))

    related: list[dict] = []
    searched_roots: set[str] = set()

    # Per query, build TWO global lists (across all indexes): the dense list,
    # cosine-sorted (cosine is comparable across indexes, so this is a true
    # global ranking — no per-index rank democracy), and the keyword/BM25 list.
    # Fuse those two by reciprocal rank: dense order is preserved, and a keyword
    # hit (e.g. a filename match with near-zero cosine) is promoted in instead of
    # being lost under the cosine floor. Then fuse across queries the same way.
    def _hybrid_for_query(q: str, want_related: bool) -> list[dict]:
        dense: list[dict] = []
        kw: list[dict] = []
        ent: list[dict] = []
        for p, root in valid_roots:
            try:
                res = query(p, q, k=k)
            except Exception as e:  # noqa: BLE001 — one bad index shouldn't break the turn
                log.warning("Grounding query failed for %s: %s", p, e)
                continue
            searched_roots.add(root)
            for m in res.get("matches", []):
                m = dict(m)
                m["_abs"] = os.path.join(root, m["path"])
                m["_ws"] = p
                dense.append(m)
            for m in res.get("keyword", []):
                m = dict(m)
                m["_abs"] = os.path.join(root, m["path"])
                m["_ws"] = p
                kw.append(m)
            try:  # entity-linking leg: chunks anchored to salient entities named in q
                for m in salience.entity_leg(p, q, k=k):
                    m = dict(m)
                    m["_abs"] = os.path.join(root, m["path"])
                    m["_ws"] = p
                    ent.append(m)
            except Exception as e:  # noqa: BLE001 — best-effort, never break the turn
                log.warning("Entity-link leg failed for %s: %s", p, e)
            if want_related:
                for r in res.get("related", []):
                    r = dict(r)
                    r["_abs"] = os.path.join(root, r["path"])
                    related.append(r)
        # Global dense ranking: cosine-sorted, sub-floor noise dropped. The floor
        # is the max of the absolute per-backend floor and a relative guard
        # (GROUND_REL_FLOOR x the top score), which adapts to per-query spread so
        # a weak tail below half the best match's relevance can't dilute grounding.
        dense.sort(key=lambda m: m.get("score", 0.0), reverse=True)
        top = dense[0].get("score", 0.0) if dense else 0.0
        floor = max(ground_floor, GROUND_REL_FLOOR * top) if top > 0 else ground_floor
        dense = [m for m in dense if m.get("score", 0.0) >= floor]
        # Keyword (BM25, more-negative = better) and entity legs are exempt from the
        # cosine floor — surfacing exact-term and entity-anchored hits is their job.
        kw.sort(key=lambda m: m.get("_bm25", 0.0))
        ent.sort(key=lambda m: m.get("_ent", 0.0), reverse=True)
        legs = [lst for lst in (dense, kw, ent) if lst]
        if len(legs) > 1:
            return _rrf_fuse(legs)
        return legs[0] if legs else []

    per_query = [_hybrid_for_query(q, want_related=(qi == 0)) for qi, q in enumerate(queries)]
    folders_searched = len(searched_roots)

    per_query = [lst for lst in per_query if lst]
    if not per_query:
        matches: list[dict] = []
    elif len(per_query) == 1:
        matches = per_query[0]
    else:
        matches = _rrf_fuse(per_query)

    # Optional cross-encoder rerank (behind the rag_reranker flag): reorder the
    # top fused candidates by judging each (question, passage) pair directly,
    # BEFORE the per-doc cap + top-k cut. The passage carries the filename so a
    # keyword/filename hit isn't demoted by content the reranker can't tie to the
    # query. Best-effort: keeps the fused order if the model is absent or fails.
    if matches:
        from server.infrastructure.feature_flags import is_enabled

        if is_enabled("rag_reranker"):
            from server.index import reranker

            head = matches[:RERANK_CANDIDATES]
            passages = [
                f"File: {os.path.basename(m.get('path', ''))}\n{m.get('text', '')}" for m in head
            ]
            scores = reranker.rerank(question, passages)
            if scores and len(scores) == len(head):
                head = [
                    m
                    for m, _ in sorted(
                        zip(head, scores, strict=False), key=lambda x: x[1], reverse=True
                    )
                ]
                matches = head + matches[RERANK_CANDIDATES:]

    # Collapse duplicate passages (byte-identical text) and cap per-document so
    # one large file can't monopolise the slots; preserves the (fused) rank order.
    deduped: list[dict] = []
    seen_text: set[str] = set()
    per_doc: dict[str, int] = {}
    for m in matches:
        tk = _text_key(m)
        if tk in seen_text:
            continue
        doc = m.get("_abs") or m.get("path")
        if per_doc.get(doc, 0) >= _GROUND_PER_DOC_CAP:
            continue
        seen_text.add(tk)
        per_doc[doc] = per_doc.get(doc, 0) + 1
        deduped.append(m)
        if len(deduped) >= k:
            break
    matches = deduped

    if not matches:
        return ("", {"folders": folders_searched, "passages": 0}) if return_meta else ""

    # Auto-merge (parent-document retrieval): when a SMALL, focused document has
    # a chunk among the top matches, splice in its remaining chunks right after
    # that chunk — so a CV's full employment history (not just its header) or a
    # contract's full terms ground the answer instead of forcing a file read.
    # Large documents (books) are gated out, and a global budget bounds growth.
    present = {_key(m) for m in matches}
    expanded_docs: set[str] = set()
    budget = _AUTOMERGE_BUDGET
    merged: list[dict] = []
    for m in matches:
        merged.append(m)
        doc_abs, ws = m.get("_abs"), m.get("_ws")
        if not doc_abs or not ws or doc_abs in expanded_docs or budget <= 0:
            continue
        expanded_docs.add(doc_abs)
        try:
            siblings = store.chunks_for_file(ws, m["path"])
        except Exception as e:  # noqa: BLE001 — auto-merge is best-effort
            log.warning("Auto-merge failed for %s: %s", m.get("path"), e)
            continue
        if len(siblings) > _AUTOMERGE_MAX_DOC_CHUNKS:
            continue  # large document (e.g. a book) — never whole-merge
        root = os.path.normpath(os.path.abspath(os.path.expanduser(ws)))
        for s in sorted(siblings, key=lambda x: x.get("start_line", 0)):
            if budget <= 0:
                break
            s = dict(s)
            s["_abs"] = os.path.join(root, s["path"])
            s["_ws"] = ws
            if _key(s) in present:
                continue
            present.add(_key(s))
            seen_text.add(_text_key(s))  # preserve byte-identical-text dedup vs the related section
            merged.append(s)
            per_doc[doc_abs] = per_doc.get(doc_abs, 0) + 1  # keep related from re-adding this doc
            budget -= 1
    matches = merged

    # Dedup the graph-hop passages against the matches (and each other) by
    # file+line range AND by normalized text, so neither the same chunk nor a
    # byte-identical copy from a duplicate file (e.g. "invoice.pdf" and
    # "invoice (1).pdf") reappears as supporting context.
    chosen = {_key(m) for m in matches}
    related_uniq: list[dict] = []
    # Rank by the salience/IDF-weighted graph score (falling back to the raw
    # shared-entity count for un-migrated indexes) so a link through a rare,
    # discriminative entity outranks one through a common word.
    for r in sorted(
        related,
        key=lambda r: (r.get("graph_score", 0.0), r.get("shared_entities", 0)),
        reverse=True,
    ):
        rk = _key(r)
        tk = _text_key(r)
        if rk in chosen or tk in seen_text:
            continue
        # Share the per-document budget with matches: a large book whose chunks
        # are entity-linked to each other must not flood the related section.
        doc = r.get("_abs") or r.get("path")
        if per_doc.get(doc, 0) >= _GROUND_PER_DOC_CAP:
            continue
        chosen.add(rk)
        seen_text.add(tk)
        per_doc[doc] = per_doc.get(doc, 0) + 1
        related_uniq.append(r)
    related_uniq = related_uniq[: max(2, k // 2)]

    def _fmt(m: dict, max_chars: int) -> str:
        link = citation_link(m["path"], m["start_line"], m["end_line"], m["_abs"])
        snippet = " ".join((m.get("text") or "").split())[:max_chars]
        return f"{link}\n   {snippet}"

    out = [
        "[Workspace index context for this question. These passages were "
        "retrieved from a semantic index of the user's files. Treat them as the "
        "source of truth and answer directly from them, citing the source links "
        "you rely on by copying them exactly as given (including the "
        "#wsfile=...&L=start-end fragment, which opens the file at the cited "
        "lines). Always end your answer with a 'Sources' section: a short list "
        "with one such link per line for every file you drew on. Only fall back "
        "to other tools if these passages are genuinely insufficient to answer; "
        "do not re-read files the index already covers.]",
        "",
    ]
    for i, m in enumerate(matches, 1):
        out.append(f"{i}. {_fmt(m, _GROUND_MATCH_CHARS)}")
    if related_uniq:
        out.append("")
        out.append("Related passages (linked through shared entities):")
        for m in related_uniq:
            out.append(f"- {_fmt(m, _GROUND_RELATED_CHARS)}")
    text = "\n".join(out)
    if return_meta:
        return text, {"folders": folders_searched, "passages": len(matches)}
    return text
